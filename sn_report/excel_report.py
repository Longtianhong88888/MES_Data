"""图片导出 Excel:行=站位,列=SN,每张图占一行,多图合并站位行。"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PilImage


ROW_HEIGHT = 120          # 每张图的行高(px)
COL_WIDTH = 30            # 列宽(字符)
IMG_WIDTH = 100           # 插入图片宽度(px),高度按比例
IMG_MARGIN = 3            # 图片间距


def _collect_rows(per_sn: Dict[str, Dict[str, Any]]) -> Tuple[List[str], List[str]]:
    """返回 (站顺序, SN顺序)。"""
    sns = list(per_sn.keys())
    stations: List[str] = []
    for sn in sns:
        for part in ("eol", "fol"):
            for img in per_sn[sn].get(part, []):
                st = img.get("station") or "?"
                label = f"{st}({part.upper()})"
                if label not in stations:
                    stations.append(label)
    return stations, sns


def _station_imgs(per_sn: Dict[str, Any], station: str, sn: str) -> List[Dict[str, Any]]:
    """某站位在某 SN 下的图片(带 part 标记)。"""
    imgs = []
    for part in ("eol", "fol"):
        for img in per_sn.get(part, []):
            label = f"{img.get('station') or '?'}({part.upper()})"
            if label == station:
                imgs.append(img)
    return imgs


def _place_image(ws, img: Dict[str, Any], row: int, col: int) -> None:
    """把图片插入单元格 (row, col),按行高 120 缩放,顶部对齐。"""
    path = img.get("dest") or ""
    if not path or not Path(path).exists():
        return
    try:
        with PilImage.open(path) as p:
            w, h = p.size
    except Exception:
        w, h = 200, 120
    if h <= 0:
        return
    target_h = ROW_HEIGHT - 2 * IMG_MARGIN
    scale = target_h / h
    iw = int(w * scale)
    if iw <= 0:
        return
    xl = XlImage(path)
    xl.width = iw
    xl.height = int(h * scale)
    # 单元格左上角锚点
    xl.anchor = f"{get_column_letter(col)}{row}"
    ws.add_image(xl)


def build_excel(per_sn: Dict[str, Dict[str, Any]], project: str,
                out_path: Path, date_str: str) -> Path:
    """生成 Excel:行=站位,列=SN。

    per_sn: {sn: {"eol": [img...], "fol": [img...]}}, img 含 station/url/dest
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    stations, sns = _collect_rows(per_sn)
    wb = Workbook()
    ws = wb.active
    ws.title = "图片"

    # 表头: A1=站位, B1..=SN
    header_fill = PatternFill("solid", fgColor="F2F2F7")
    header_font = Font(bold=True, color="1D1D1F")
    thin = Side(style="thin", color="E5E5EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=1, value="站位")
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    for i, sn in enumerate(sns, start=2):
        c = ws.cell(row=1, column=i, value=sn)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.cell(row=1, column=1).border = border

    # 列宽
    ws.column_dimensions["A"].width = COL_WIDTH
    for i in range(2, len(sns) + 2):
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTH

    # 每个站位占 n 行(n=该站位所有 SN 图片数的最大值),行合并
    current_row = 2
    for st in stations:
        max_n = 1
        for sn in sns:
            n = len(_station_imgs(per_sn[sn], st, sn))
            max_n = max(max_n, max(1, n))
        # 站位名合并 max_n 行
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row + max_n - 1, end_column=1)
        c = ws.cell(row=current_row, column=1, value=st)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        for r in range(current_row, current_row + max_n):
            ws.row_dimensions[r].height = ROW_HEIGHT
            ws.cell(row=r, column=1).border = border
            for i in range(2, len(sns) + 2):
                ws.cell(row=r, column=i).border = border

        # 各 SN 在该站位的图片:每张图占一行(单元格),站位名列已合并
        for si, sn in enumerate(sns, start=2):
            imgs = _station_imgs(per_sn[sn], st, sn)
            # 每张图对应一行,从 current_row 开始顺排
            for k, img in enumerate(imgs[:max_n]):
                r = current_row + k
                _place_image(ws, img, r, si)

        current_row += max_n

    wb.save(str(out_path))
    return out_path
