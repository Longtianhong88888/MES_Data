"""配置读取:根目录 config.json(账号/入口)+ sn_report/config.json(功能配置)。"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any, Dict, List


def _find_sn_report_dir() -> Path:
    """定位 sn_report 目录:打包成 exe 后 __file__ 指向临时目录,
    优先用当前工作目录下的 sn_report(与项目目录共存),其次用模块路径。"""
    cwd_candidate = Path.cwd() / "sn_report"
    if cwd_candidate.is_dir():
        return cwd_candidate
    return Path(__file__).resolve().parent.parent


SN_REPORT_DIR = _find_sn_report_dir()
PROJECT_DIR = SN_REPORT_DIR.parent

DEFAULT_SN_REPORT_CONFIG: Dict[str, Any] = {
    "sn_list_file": "sns.txt",
    "analysis_window": {"start": "2026-06-01 00:00", "end": "2026-08-08 23:59"},
    "report": {
        "output_dir": "output",
        "download_dir": "downloads",
        "ppt_name": "SN全制程追溯报告.pptx",
        "slide_font": "Microsoft YaHei",
        "download_images": False,
    },
    "sn_search": {"page": "report/snsearch.aspx", "field": "sntextbox", "button": "Button1"},
    "sntotalinfo_page": "Tracking/sntotalinfo.aspx",
    "img_stations": [
        {"id": "cubepnpimguploadbak", "label": "CUBEPNP"},
        {"id": "aaimguploadbak", "label": "AA"},
        {"id": "LMimguploadbak", "label": "LM"},
        {"id": "aviimguploadbak", "label": "AVI"},
        {"id": "frtopimguploadbak", "label": "FRTOP"},
        {"id": "acfflipimguploadbak", "label": "ACFFlip"},
    ],
    "acf_mc_types": [
        {"id": "acfbondnewdatabak", "label": "ACF上料機"},
        {"id": "acfunloadbondnewdatabak", "label": "ACF下料機"},
        {"id": "acfmaindatanewbak", "label": "ACF主機"},
    ],
    "column_keywords": {
        "station": ["站位", "工序", "站点", "站點", "工序名"],
        "time": ["进站时间", "進站時間", "开始时间", "開始時間", "Start_Time", "start_time", "时间", "時間"],
        "mc": ["机台", "機台", "设备", "設備", "MC_ID", "Mc_Id", "machine"],
        "carrier": ["载板", "載板", "Carrier_ID", "carrier_id", "carrier"],
        "pocket": ["穴位", "Cavity", "cavity", "Pocket", "pocket"],
        "head": ["Head_ID", "head_id", "Head"],
        "material": ["材料", "物料", "耗材", "名称", "名稱", "Lot", "lot", "批号", "批號", "LOT_ID"],
        "summary": ["批号", "批號", "线体", "線體", "包号", "包號", "SFC", "测试结果", "測試結果", "包装", "包裝", "箱号", "箱號"],
    },
    "c4": {
        "enabled": False,
        "url": "http://10.151.128.35:8095/api/MachineParameter/GetInformationDT",
        "plant_id": "",
        "device": "",
        "token": "",
        "type": "8S01",
        "extra_params": {},
        "columns": [
            # {"station": "OIS PNP", "mc": "", "carrier": "", "pocket": "", "start_time": ""}
        ],
    },
}


def _load_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8-sig"))


def get_sn_report_config() -> Dict[str, Any]:
    """读取 sn_report/config.json;不存在时写出默认配置并返回。"""
    path = SN_REPORT_DIR / "config.json"
    cfg = _load_json(path)
    merged = DEFAULT_SN_REPORT_CONFIG.copy()
    _deep_merge(merged, cfg)
    if not path.exists():
        path.write_text(
            json.dumps(DEFAULT_SN_REPORT_CONFIG, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"[config] 已生成默认配置: {path} (请按需修改)")
    return merged


def _deep_merge(base: Dict[str, Any], override: Dict[str, Any]) -> None:
    for k, v in override.items():
        if isinstance(v, dict) and isinstance(base.get(k), dict):
            _deep_merge(base[k], v)
        else:
            base[k] = v


def get_root_config() -> Dict[str, Any]:
    """读取项目根目录 config.json(登录账号 / 入口 URL)。"""
    path = PROJECT_DIR / "config.json"
    if not path.exists():
        raise FileNotFoundError(
            f"缺少项目根目录 config.json: {path}\n请复制 config.example.json 并填写账号信息。"
        )
    return _load_json(path)


def load_sn_list(sn_list_path: Path) -> List[str]:
    """读取 SN 列表,支持 .txt(每行一个)/ .csv / .xlsx。"""
    sn_list_path = Path(sn_list_path)
    if not sn_list_path.exists():
        raise FileNotFoundError(f"SN 列表不存在: {sn_list_path}")

    ext = sn_list_path.suffix.lower()
    sns: List[str] = []

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - Windows 环境缺依赖时提示
            raise RuntimeError("读取 xlsx 需要 openpyxl,请先安装(见 sn_report/README.md)") from exc
        wb = load_workbook(sn_list_path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        for r_idx, row in enumerate(rows):
            if r_idx == 0 and row and str(row[0]).strip().lower() in ("serial_no", "sn", "serial"):
                continue
            if row and row[0] is not None:
                sn = str(row[0]).strip()
                if sn:
                    sns.append(sn)
        wb.close()
    elif ext == ".csv":
        import csv

        with open(sn_list_path, newline="", encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                if not row:
                    continue
                sn = row[0].strip()
                if sn and sn.lower() not in ("serial_no", "sn"):
                    sns.append(sn)
    else:
        for line in sn_list_path.read_text(encoding="utf-8-sig").splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                sns.append(line)

    seen = set()
    deduped = []
    for sn in sns:
        if sn not in seen:
            seen.add(sn)
            deduped.append(sn)
    return deduped


def ensure_windows_lib() -> None:
    """Windows 下优先使用项目内 lib/ 离线依赖(免安装)。"""
    if not sys.platform.startswith("win"):
        return
    lib_dir = PROJECT_DIR / "lib"
    if (lib_dir / "requests").is_dir():
        sys.path.insert(0, str(lib_dir))
