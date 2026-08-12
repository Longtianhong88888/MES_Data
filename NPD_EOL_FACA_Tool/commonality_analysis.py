#!/usr/bin/env python3
"""共性分析:单因素评分排行(替代纯桑基图目视)。

对每个维度(MC/批号/Vendor/Carrier/Head/穴位/胶水…)× 每个值,统一计算:
  - fail_count / fail_rate / lift / fail_ratio
  - Fisher 精确检验 p 值(手写,无 scipy 依赖)+ Benjamini-Hochberg FDR 校正
  - score = fail_ratio * log2(lift),再过滤 min_fail 与显著性
输出 Top N 共性可疑点表(CSV/Excel)。

用法:
    python commonality_analysis.py --csv <共性CSV> --out output/共性Top20.xlsx
    python commonality_analysis.py --csv ... --fail-col Result --key-col Serial_No
"""

from __future__ import annotations

import argparse
import itertools
import math
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd


# ---------- Fisher 精确检验(2x2, 双侧, 无 scipy) ----------
def _log_comb(n: int, k: int) -> float:
    if k < 0 or k > n:
        return float("-inf")
    return (math.lgamma(n + 1) - math.lgamma(k + 1)
            - math.lgamma(n - k + 1))


def fisher_exact_2x2(a: int, b: int, c: int, d: int) -> float:
    """2x2 表 [[a,b],[c,d]],返回双侧 p 值(精确超几何)。"""
    n = a + b + c + d
    row1, row2 = a + b, c + d
    col1, col2 = a + c, b + d
    if row1 == 0 or row2 == 0 or col1 == 0 or col2 == 0:
        return 1.0
    k_min = max(0, col1 - row2)
    k_max = min(row1, col1)
    log_p_obs = (_log_comb(row1, a) + _log_comb(row2, c)
                 - _log_comb(n, col1))
    p_obs = math.exp(log_p_obs) if log_p_obs > -745 else 0.0
    p = 0.0
    for k in range(k_min, k_max + 1):
        lp = (_log_comb(row1, k) + _log_comb(row2, col1 - k)
              - _log_comb(n, col1))
        pk = math.exp(lp) if lp > -745 else 0.0
        if pk <= p_obs * 1.0001:
            p += pk
    return min(1.0, p)


def bh_fdr(pvals: List[float]) -> List[float]:
    """Benjamini-Hochberg FDR 校正。"""
    m = len(pvals)
    if m == 0:
        return []
    order = sorted(range(m), key=lambda i: pvals[i])
    ranked = [0.0] * m
    prev = 1.0
    for rank, idx in enumerate(reversed(order), start=1):
        adj = pvals[idx] * m / rank
        adj = min(prev, adj)
        ranked[idx] = min(1.0, adj)
        prev = adj
    return ranked


# ---------- 因子列识别 ----------
_INCLUDE = re.compile(
    r"(?i)(MC_ID|Lot_ID|Vendor|Carrier_ID|Pocket|Head_ID|Syringe_ID|"
    r"Shuttleno|Driver_UID|Lorentz_Version|Part_SN|Tool_ID|Cavity_ID|"
    r"Platform|Film_Lot|Wafer_ID|wafer_pocket|APN|Bump|solder_balls_ID|"
    r"EXY_Syringe_ID)")
_EXCLUDE = re.compile(
    r"(?i)(Time|Date|Usage_Time|Expired|Staging|Serial_No|Result|"
    r"Failed_Station|Failure_Mode|Project|Rev|Site|Build|Config|"
    r"XY$|Shuttlepos|Part_SN)")

# 六大类共性图表分类(与共性报告 PPT 6 类图一致)
CHART_MACHINE = "machine"
CHART_MATERIAL = "material"
CHART_HEAD = "head"
CHART_TIME = "time"
CHART_WAFER = "wafer"
CHART_CARRIER = "carrier"
CATEGORY_NAMES = {
    CHART_MACHINE: "机台",
    CHART_MATERIAL: "材料",
    CHART_HEAD: "Head",
    CHART_TIME: "时间",
    CHART_WAFER: "Wafer",
    CHART_CARRIER: "穴位",
}
CATEGORY_ORDER = [CHART_MACHINE, CHART_MATERIAL, CHART_HEAD,
                  CHART_TIME, CHART_WAFER, CHART_CARRIER]


def chart_category(dim: str) -> str:
    """把维度列名归到 6 大类(机台/材料/Head/时间/Wafer/穴位),否则其他。"""
    l = str(dim).lower()
    if l.endswith("_mc_id"):
        return CHART_MACHINE
    if l.endswith("_head_id") or "_head_id" in l:
        return CHART_HEAD
    if l.endswith("_start_time") or "_time" in l:
        return CHART_TIME
    if "carrier_pocket" in l or "pocket" in l or "cavity" in l:
        return CHART_CARRIER
    if "wafer" in l and ("xy" in l or "_x" in l or "_y" in l):
        return CHART_WAFER
    if ("lot_id" in l or "vendor" in l or "material" in l
            or "wafer" in l):
        return CHART_MATERIAL
    return "其他"


def dimension_type(col: str) -> str:
    """给因子列分类:NPI 重点关注的 材料/治工具/工作区/机台。"""
    c = col.upper()
    if c.endswith("MC_ID") or c.endswith("STATION"):
        return "机台"
    if ("HEAD_ID" in c or "CARRIER_POCKET" in c or "GLUE_PLATFORM" in c
            or "PLATFORM" in c):
        return "工作区"
    if ("TOOL_ID" in c or "CAVITY_ID" in c or "SHUTTLE" in c
            or "DRIVER_UID" in c or "LORENTZ" in c):
        return "治工具"
    if ("VENDOR" in c or "LOT_ID" in c or "PART_SN" in c or "APN" in c
            or "FILM" in c or "SOLDER" in c or "SYRINGE" in c):
        return "材料/耗材"
    return "其他"


def pick_factor_columns(cols: List[str]) -> List[str]:
    out = []
    for c in cols:
        if _EXCLUDE.search(c):
            continue
        if _INCLUDE.search(c):
            out.append(c)
    return out


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s in ("{}", "nan", "None", ""):
        return ""
    return s


def analyze_commonality(
    df: pd.DataFrame,
    key_col: str = "Serial_No",
    fail_col: str = "Result",
    fail_values: Tuple[str, ...] = ("Fail",),
    min_fail: int = 3,
    factor_cols: Optional[List[str]] = None,
    coverage_penalty: float = 1.0,
    coverage_threshold: float = 0.8,
    mode: str = "mp",
) -> Tuple[List[Dict], pd.DataFrame]:
    """返回 (rows, factor_df)。rows 按 score 降序的共性可疑点。

    coverage_penalty: 覆盖率降权指数(仅当 fail_ratio > coverage_threshold 时生效)。
        score = fail_ratio * log2(lift) * penalty,其中
        penalty = 1(fail_ratio<=threshold) 或
        ((1-fail_ratio)/(1-threshold))^coverage_penalty(>threshold,线性衰减到 0)。
        默认 threshold=0.8:覆盖 80% 以下不罚(UTAC 76.7% 保留),100% 覆盖(如
        SONY 全覆盖)被压掉。
    mode: "mp"=量产(默认,min_fail=3+FDR);"npi"=试产小样本(min_fail=1、跳过 FDR、
        突出 Fail 共享覆盖);"auto"=按 Fail 数自动(<20 走 npi)。
    """
    if factor_cols is None:
        factor_cols = pick_factor_columns(list(df.columns))
    fail_col = fail_col if fail_col in df.columns else "pass_fail"
    if fail_col not in df.columns:
        raise ValueError("找不到 pass/fail 列: %s" % fail_col)

    is_fail = df[fail_col].astype(str).str.strip().isin(fail_values)
    total = int(len(df))
    total_fail = int(is_fail.sum())
    total_pass = total - total_fail
    overall_rate = total_fail / total if total else 0.0
    if mode == "auto":
        mode = "npi" if total_fail < 20 else "mp"
    if mode == "npi":
        min_fail = max(1, min_fail)
    if total_fail < min_fail:
        return [], df[factor_cols]

    rows: List[Dict] = []
    for col in factor_cols:
        vals = df[col].map(_clean)
        for value, group in vals.groupby(vals):
            if not value:
                continue
            n_total = int(len(group))
            n_fail = int(is_fail.loc[group.index].sum())
            if n_fail < min_fail:
                continue
            n_pass = n_total - n_fail
            rate = n_fail / n_total if n_total else 0.0
            lift = rate / overall_rate if overall_rate else 0.0
            ratio = n_fail / total_fail if total_fail else 0.0
            p = fisher_exact_2x2(
                n_fail, n_pass, total_fail - n_fail, total_pass - n_pass)
            rows.append({
                "dimension": col,
                "dim_type": dimension_type(col),
                "value": value,
                "fail_count": n_fail,
                "unit_count": n_total,
                "fail_rate": round(rate, 4),
                "overall_fail_rate": round(overall_rate, 4),
                "lift": round(lift, 2),
                "fail_ratio": round(ratio, 4),
                "p_value": p,
            })

    if not rows:
        return [], df[factor_cols]
    if mode == "npi":
        # NPI 小样本:不做 FDR(太保守),p_adj 用原始 p,并在报告注明
        for r in rows:
            r["p_adj"] = r["p_value"]
            r["score"] = round(
                r["fail_ratio"] * math.log2(max(r["lift"], 1.0)), 4)
    else:
        pvals = [r["p_value"] for r in rows]
        padjs = bh_fdr(pvals)
        for r, pa in zip(rows, padjs):
            r["p_adj"] = pa
            if r["fail_ratio"] <= coverage_threshold:
                pen = 1.0
            else:
                denom = 1.0 - coverage_threshold
                pen = ((1.0 - r["fail_ratio"]) / denom) ** coverage_penalty \
                    if denom > 0 else 0.0
            r["score"] = round(
                r["fail_ratio"] * math.log2(max(r["lift"], 1.0))
                * max(0.0, pen), 4)
    rows.sort(key=lambda r: r["score"], reverse=True)
    return rows, df[factor_cols]


def apriori_rules(
    df: pd.DataFrame,
    fail_col: str = "pass_fail",
    min_support: float = 0.15,
    min_lift: float = 1.5,
    top_n: int = 20,
    max_items: int = 40,
    drop_structural: bool = True,
) -> pd.DataFrame:
    """第二层:在 Fail 子集上挖掘 2-项组合规则(Apriori 精简版)。

    找出 Fail 中频繁共现的"维度=值"组合(如 UTAC 加强片 + COT 胶),
    按 lift 排序 —— 抓单因素分析发现不了的组合共性。
    drop_structural: 过滤"机器×机器"的结构性共现。每台货必然经过每个站位的
    一台机,所以两个 MC_ID 组合天然共现(如 EA0403+FC0403),非新信息;
    保留跨维度真组合(如 Vendor×Vendor / Vendor×机器)。
    """
    is_fail = df[fail_col].astype(str).str.strip().isin(("fail", "Fail"))
    fail_df = df[is_fail]
    n = int(len(fail_df))
    if n == 0:
        return pd.DataFrame()
    factor_cols = pick_factor_columns(list(df.columns))
    rowsets = []
    for _, r in fail_df.iterrows():
        items = []
        for col in factor_cols:
            v = r[col]
            v = _clean(v)
            if v:
                items.append(f"{col}={v}")
        rowsets.append(set(items))
    one = Counter()
    for s in rowsets:
        one.update(s)
    min_cnt = max(3, int(n * min_support))
    freq1 = {k: c for k, c in one.items() if c >= min_cnt}
    top_items = sorted(freq1, key=freq1.get, reverse=True)[:max_items]
    if len(top_items) < 2:
        return pd.DataFrame()
    pair = Counter()
    for s in rowsets:
        s2 = s & set(top_items)
        if len(s2) >= 2:
            for a, b in itertools.combinations(sorted(s2), 2):
                pair[(a, b)] += 1
    rules = []
    for (a, b), ab in pair.items():
        if ab < min_cnt:
            continue
        ca, cb = freq1[a], freq1[b]
        p_a, p_b, p_ab = ca / n, cb / n, ab / n
        lift = p_ab / (p_a * p_b) if p_a and p_b else 0.0
        if lift >= min_lift:
            structural = False
            if drop_structural:
                a_dim, a_val = a.split("=", 1)
                b_dim, b_val = b.split("=", 1)
                if a_dim.endswith("MC_ID") and b_dim.endswith("MC_ID"):
                    structural = True
            rules.append({
                "item_a": a, "item_b": b,
                "count_a": ca, "count_b": cb, "count_ab": ab,
                "support": round(p_ab, 4),
                "lift": round(lift, 2),
                "structural": structural,
            })
    rules.sort(key=lambda r: r["lift"], reverse=True)
    if drop_structural:
        rules = [r for r in rules if not r["structural"]]
    out = pd.DataFrame(rules).head(top_n)
    if not out.empty:
        out = out[["item_a", "item_b", "count_ab", "support", "lift"]]
    return out


def _line_no(value: str) -> str:
    """从机台值提取线号,如 CA0402->0402, OP1-0403->0403, UFC0403-2->0403。"""
    m = re.search(r"(\d{4})(?:-\d+)?$", value)
    return m.group(1) if m else ""


# ---------- Excel 样式输出(重点/次重点配色 + 栏宽自适应) ----------
def write_styled_excel(
    path,
    top_df: pd.DataFrame,
    rules_df: Optional[pd.DataFrame] = None,
    score_key: float = 1.5,
    score_second: float = 0.8,
    lift_key: float = 2.5,
    lift_second: float = 2.0,
    mode: str = "mp",
) -> None:
    """写 Excel:Top20 按 score 配色(重点红/次重点橙),组合规则按 lift 配色;
    栏宽随内容自适应;数据右侧附"阅读说明"。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    fill_key = PatternFill("solid", fgColor="FFC7CE")      # 红:重点
    fill_second = PatternFill("solid", fgColor="FFEB9C")   # 橙:次重点
    fill_hdr = PatternFill("solid", fgColor="DDEBF7")
    font_hdr = Font(bold=True)
    font_key = Font(color="9C0006", bold=True)
    font_second = Font(color="9C6500")

    GUIDES = {
        "Top20_mp": [
            "【阅读说明】",
            "· 类别列:按 机台/材料/Head/时间/Wafer/穴位 6 大类分组",
            "· Score = Fail占比 × log2(Lift)，越高越可疑",
            "· 红底=重点(Score≥1.5)；橙底=次重点(Score≥0.8)",
            "· Lift：该值不良率÷整体不良率，>1 即异常放大",
            "· Fail占比：该值Fail数÷总Fail数(覆盖率)",
            "· Fail率：该值Fail数÷该值样本数",
            "· p_adj：FDR校正后显著性，<0.05 才可信",
            "· Fail数<3 不参与统计(防噪声)",
            "· 样本数含Pass采样(默认3000/天)，为基线口径",
            "· 覆盖>80%且Lift低的项被降权(必要非充分条件)",
        ],
        "Top20_npi": [
            "【阅读说明·NPI试产小样本】",
            "· 类别列:按 机台/材料/Head/时间/Wafer/穴位 6 大类分组",
            "· Fail数少(<20)，不适用FDR，直接看原始p值",
            "· Score = Fail占比 × log2(Lift)，越高越可疑",
            "· 红底=重点(Score≥1.5)；橙底=次重点(Score≥0.8)",
            "· 重点看 Fail 共用的:材料/治工具/工作区(dim_type列)",
            "· Lift：该值Fail比例÷整体，>1 即异常",
            "· Fail占比：NPI Fail中有多少共用该值(覆盖率)",
            "· p值：与Pass基线对比的Fisher精确检验，越小越异常",
            "· 样本小，结论当线索；建议人工复核机台内部工作区",
        ],
        "组合规则_mp": [
            "【阅读说明】",
            "· Lift：两项在Fail中共现频率÷随机期望，>1.5 有意义",
            "· 红底=重点(Lift≥2.5)；橙底=次重点(Lift≥2.0)",
            "· count_ab：两项共同出现的Fail数",
            "· support：共现Fail数÷总Fail数",
            "· 已过滤“机器×机器”结构性共现",
            "· 跨维度组合(如供应商×供应商)是真共性线索",
        ],
        "组合规则_npi": [
            "【阅读说明·NPI试产】",
            "· 小样本下组合规则仅作参考",
            "· Lift：两项共现÷随机期望，>1.5 值得关注",
            "· count_ab：两项共同出现的Fail数",
            "· 已过滤“机器×机器”结构性共现",
        ],
    }

    def write_sheet(ws, df, value_col, th_key, th_second, guide_key):
        ws.append(list(df.columns))
        for c in ws[1]:
            c.fill = fill_hdr
            c.font = font_hdr
            c.alignment = Alignment(horizontal="center")
        for _, row in df.iterrows():
            ws.append([row[c] for c in df.columns])
            r = ws.max_row
            v = float(row[value_col])
            if v >= th_key:
                for c in ws[r]:
                    c.fill = fill_key
                    c.font = font_key
            elif v >= th_second:
                for c in ws[r]:
                    c.fill = fill_second
                    c.font = font_second
        # 栏宽自适应
        for col in ws.columns:
            letter = col[0].column_letter
            max_len = max(len(str(c.value)) for c in col if c.value is not None)
            ws.column_dimensions[letter].width = min(max(max_len + 4, 10), 52)
        # 数据右侧附阅读说明
        gcol = len(df.columns) + 2
        gl = get_column_letter(gcol)
        ws.cell(row=1, column=gcol, value="阅读说明")
        ws.cell(row=1, column=gcol).font = Font(bold=True)
        ws.cell(row=1, column=gcol).fill = fill_hdr
        for i, line in enumerate(GUIDES.get(guide_key, []), start=2):
            cell = ws.cell(row=i, column=gcol, value=line)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if line.startswith("【"):
                cell.font = Font(bold=True)
        ws.column_dimensions[gl].width = 46

    # Top20 加"类别"列并按 6 大类排序(机台→材料→Head→时间→Wafer→穴位→其他)
    top_df = top_df.copy()
    top_df["类别"] = top_df["dimension"].map(
        lambda d: CATEGORY_NAMES.get(chart_category(d), "其他"))
    _cat_order = {CATEGORY_NAMES[k]: i for i, k in enumerate(CATEGORY_ORDER)}
    _cat_order["其他"] = 99
    top_df["_cat_order"] = top_df["类别"].map(_cat_order)
    top_df = top_df.sort_values(
        ["_cat_order", "score"], ascending=[True, False]).reset_index(drop=True)
    top_df = top_df.drop(columns=["_cat_order"])
    # 类别列放在"维度"之后
    _cols = list(top_df.columns)
    if "dimension" in _cols and "类别" in _cols:
        _cols.remove("类别")
        _cols.insert(_cols.index("dimension") + 1, "类别")
        top_df = top_df[_cols]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Top20"
    guide1 = "Top20_npi" if mode == "npi" else "Top20_mp"
    write_sheet(ws1, top_df, "score", score_key, score_second, guide1)
    if rules_df is not None and not rules_df.empty:
        ws2 = wb.create_sheet("组合规则")
        guide2 = "组合规则_npi" if mode == "npi" else "组合规则_mp"
        write_sheet(ws2, rules_df, "lift", lift_key, lift_second, guide2)
    wb.save(path)


def load_csv(path: Path, factor_cols: Optional[List[str]] = None) -> pd.DataFrame:
    """读取共性 CSV(可只读需要的列)。"""
    usecols = None
    if factor_cols:
        usecols = list(factor_cols)
    # 先探测表头
    hdr = pd.read_csv(path, nrows=0, encoding="utf-8-sig").columns.tolist()
    if usecols is None:
        usecols = ["Serial_No", "Result"] + pick_factor_columns(hdr)
    usecols = [c for c in usecols if c in hdr]
    if "Result" in hdr and "Result" not in usecols:
        usecols.append("Result")
    if "Serial_No" in hdr and "Serial_No" not in usecols:
        usecols.append("Serial_No")
    return pd.read_csv(path, usecols=usecols, low_memory=False,
                       encoding="utf-8-sig", dtype=str)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", required=True, help="共性数据 CSV(含 Result/Serial_No)")
    ap.add_argument("--out", default="output/commonality_top20.xlsx")
    ap.add_argument("--min-fail", type=int, default=3)
    ap.add_argument("--top", type=int, default=20)
    ap.add_argument("--fail-col", default="Result")
    ap.add_argument("--key-col", default="Serial_No")
    ap.add_argument("--coverage-penalty", type=float, default=1.0)
    ap.add_argument("--rules", action="store_true", help="输出第二层组合规则")
    ap.add_argument("--min-support", type=float, default=0.15)
    ap.add_argument("--min-lift", type=float, default=1.5)
    ap.add_argument("--mode", default="mp", choices=["mp", "npi", "auto"],
                    help="mp=量产(默认);npi=试产小样本;auto=按Fail数自动")
    args = ap.parse_args()

    csv_path = Path(args.csv)
    print("读取 %s ..." % csv_path, flush=True)
    df = load_csv(csv_path)
    print("行数: %d, 列数: %d" % (len(df), len(df.columns)), flush=True)
    rows, _ = analyze_commonality(
        df, key_col=args.key_col, fail_col=args.fail_col,
        min_fail=args.min_fail, coverage_penalty=args.coverage_penalty,
        mode=args.mode)
    if not rows:
        print("无显著性共性点")
        return 0
    out = pd.DataFrame(rows).head(args.top)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    rules = pd.DataFrame()
    if args.rules:
        rules = apriori_rules(
            df, fail_col=args.fail_col,
            min_support=args.min_support, min_lift=args.min_lift)
    print("\n=== 共性可疑点 Top %d ===\n" % len(out))
    print(out.to_string(index=False))
    if out_path.suffix.lower() == ".xlsx":
        write_styled_excel(out_path, out, rules if not rules.empty else None,
                           mode=args.mode)
    else:
        out.to_csv(out_path, index=False)
    print("\n保存: %s(重点红/次重点橙,栏宽自适应)" % out_path)
    if args.rules:
        if not rules.empty:
            print("\n=== 组合规则(第二层) ===\n")
            print(rules.to_string(index=False))
        else:
            print("\n无满足条件的组合规则(min_support=%.2f, min_lift=%.1f)"
                  % (args.min_support, args.min_lift))
    return 0


if __name__ == "__main__":
    sys.exit(main())
