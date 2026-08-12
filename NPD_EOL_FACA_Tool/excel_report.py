"""图片导出 Excel:行=站位,列=SN,每个 SN 分【图片】【链接】两列。

图片列按单元格缩放嵌入(避免宽图被裁切);链接列放全部原始 URL,
点击可在浏览器打开完整原图(源图若为裁切版,用链接看原始)。
"""

from __future__ import annotations

import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PilImage


ROW_HEIGHT = 120          # 每张图的行高(px)
IMG_COL_WIDTH = 16        # 图片列宽(字符)
LINK_COL_WIDTH = 45       # 链接列宽(字符)
STATION_COL_WIDTH = 30    # 站位列宽(字符)
IMG_MARGIN = 3            # 图片间距
IMG_BOX_W = 110           # 图片嵌入最大宽度(px,适配单元格避免裁切)
IMG_BOX_H = 112           # 图片嵌入最大高度(px)


def _parse_ts(v: Any) -> Optional[datetime]:
    """解析上传时间(支持 datetime 与常见字符串格式)。"""
    if isinstance(v, datetime):
        return v
    s = str(v or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
                "%Y%m%d %H:%M:%S", "%Y%m%d%H%M%S",
                "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _collect_rows(per_sn: Dict[str, Dict[str, Any]]) -> Tuple[List[str], List[str]]:
    """返回 (站顺序, SN顺序)。"""
    sns = list(per_sn.keys())
    meta: Dict[str, Tuple[int, Optional[datetime], int]] = {}
    for sn in sns:
        # 制程顺序:FOL 在前,EOL 在后
        for part in ("fol", "eol"):
            for img in per_sn[sn].get(part, []):
                st = img.get("station") or "?"
                label = f"{st}({part.upper()})"
                t = _parse_ts(img.get("uploadtime"))
                if label not in meta:
                    meta[label] = (0 if part == "fol" else 1, t, len(meta))
                elif t and (meta[label][1] is None or t < meta[label][1]):
                    meta[label] = (meta[label][0], t, meta[label][2])
    # 分组(FOL 先,EOL 后)内按最早上传时间排序;无时间的排最后,保持出现顺序
    stations = sorted(
        meta.keys(),
        key=lambda lb: (meta[lb][0],
                        meta[lb][1] if meta[lb][1] is not None
                        else datetime.max,
                        meta[lb][2]))
    return stations, sns


def _station_imgs(per_sn: Dict[str, Any], station: str, sn: str) -> List[Dict[str, Any]]:
    """某站位在某 SN 下的图片(带 part 标记)。"""
    imgs = []
    for part in ("eol", "fol"):
        for img in per_sn.get(part, []):
            label = f"{img.get('station') or '?'}({part.upper()})"
            if label == station:
                imgs.append(img)
    return imgs


def _place_image(ws, img: Dict[str, Any], row: int, col: int) -> bool:
    """把图片转 JPEG 后插入单元格,缩放适配 [IMG_BOX_W x IMG_BOX_H] 方框,
    宽图不再超出列宽被裁切。成功返回 True,失败返回 False。"""
    path = img.get("dest") or ""
    ok = False
    if path and Path(path).exists():
        try:
            with PilImage.open(path) as p:
                w, h = p.size
                p.verify()
            ok = True
        except Exception:
            ok = False
    if ok:
        # 统一转 JPEG,避免 Excel 里出现 PNG
        try:
            tmp = Path(tempfile.gettempdir()) / (
                f"xl_{abs(hash(str(path)))}_{Path(path).stem[:10]}.jpg"
            )
            with PilImage.open(path) as p:
                p.verify()
            with PilImage.open(path) as p:
                p.convert("RGB").save(tmp, "JPEG", quality=92)
            path = tmp
        except Exception as exc:  # noqa: BLE001
            print(f"  [excel] JPEG 转换失败: {str(exc)[:60]}")
            ok = False
    if ok and h <= 0:
        ok = False
    if ok:
        box_w, box_h = IMG_BOX_W, IMG_BOX_H
        scale = min(box_w / w, box_h / h)
        iw = int(w * scale)
        ih = int(h * scale)
        if iw > 0 and ih > 0:
            xl = XlImage(str(path))
            xl.width = iw
            xl.height = ih
            xl.anchor = f"{get_column_letter(col)}{row}"
            ws.add_image(xl)
            return True
    # 图片缺失/损坏: 在单元格写入下载链接(可点击,自动换行)
    url = str(img.get("url") or "")
    if url:
        cell = ws.cell(row=row, column=col)
        cell.value = url
        cell.hyperlink = url
        cell.alignment = Alignment(
            horizontal="left", vertical="top",
            wrap_text=True, text_rotation=0,
        )
        cell.font = Font(color="007AFF", size=9)
    return False


def _place_links(ws, imgs: List[Dict[str, Any]], row: int, col: int,
                 max_n: int):
    """链接列:每张图一行,写入原始 URL 超链接(可点击打开原图)。"""
    for k in range(max_n):
        cell = ws.cell(row=row + k, column=col)
        if k < len(imgs):
            url = str(imgs[k].get("url") or "")
            if url:
                cell.value = url
                cell.hyperlink = url
        cell.alignment = Alignment(
            horizontal="left", vertical="center",
            wrap_text=True, text_rotation=0)
        cell.font = Font(color="007AFF", size=9)


def build_excel(per_sn: Dict[str, Dict[str, Any]], project: str,
                out_path: Path, date_str: str) -> Path:
    """生成 Excel:行=站位,列=SN。

    per_sn: {sn: {"eol": [img...], "fol": [img...]}}, img 含 station/url/dest
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    stations, sns = _collect_rows(per_sn)
    wb = Workbook()
    ws = wb.active
    ws.title = "图片"

    # 表头: 两行
    #   行1: 站位 | SN1(跨2列) | SN2(跨2列) | ...
    #   行2:      | 图片 | 链接 | 图片 | 链接 | ...
    header_fill = PatternFill("solid", fgColor="F2F2F7")
    header_font = Font(bold=True, color="1D1D1F")
    thin = Side(style="thin", color="E5E5EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=1, value="站位")
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=1).fill = header_fill
    for i, sn in enumerate(sns):
        col_img = 2 + i * 2
        col_link = col_img + 1
        ws.merge_cells(start_row=1, start_column=col_img,
                       end_row=1, end_column=col_link)
        c = ws.cell(row=1, column=col_img, value=sn)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c2 = ws.cell(row=2, column=col_img, value="图片")
        c2.fill = header_fill
        c2.font = header_font
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c3 = ws.cell(row=2, column=col_link, value="链接")
        c3.fill = header_fill
        c3.font = header_font
        c3.alignment = Alignment(horizontal="center", vertical="center")
        for cc in (c, c2, c3):
            cc.border = border
    ws.cell(row=1, column=1).border = border
    ws.cell(row=2, column=1).border = border
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    # 列宽
    ws.column_dimensions["A"].width = STATION_COL_WIDTH
    for i in range(len(sns)):
        ws.column_dimensions[get_column_letter(2 + i * 2)].width = IMG_COL_WIDTH
        ws.column_dimensions[get_column_letter(3 + i * 2)].width = LINK_COL_WIDTH

    # 每个站位占 n 行(n=该站位所有 SN 图片数的最大值),行合并
    current_row = 3
    for st in stations:
        max_n = 1
        for sn in sns:
            n = len(_station_imgs(per_sn[sn], st, sn))
            max_n = max(max_n, max(1, n))
        # 站位名合并 max_n 行
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row + max_n - 1, end_column=1)
        c = ws.cell(row=current_row, column=1, value=st)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        for r in range(current_row, current_row + max_n):
            ws.row_dimensions[r].height = ROW_HEIGHT
            ws.cell(row=r, column=1).border = border
            for i in range(2, len(sns) * 2 + 2):
                ws.cell(row=r, column=i).border = border

        # 各 SN 在该站位的图片:【图片】列嵌图,【链接】列放原始 URL
        for si, sn in enumerate(sns):
            col_img = 2 + si * 2
            col_link = col_img + 1
            imgs = _station_imgs(per_sn[sn], st, sn)
            for k, img in enumerate(imgs[:max_n]):
                r = current_row + k
                _place_image(ws, img, r, col_img)
            _place_links(ws, imgs, current_row, col_link, max_n)

        current_row += max_n

    wb.save(str(out_path))
    return out_path
